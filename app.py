import os
import re
import json
import logging
from datetime import datetime

from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from pymongo import MongoClient
from flask_bcrypt import Bcrypt

import PyPDF2
import requests
import openpyxl

import joblib
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# ---------------- Logging ----------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ---------------- Flask ----------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")
EXCEL_FILE_PATH = os.path.join(BASE_DIR, "jobrolespskillsframeworks.xlsx")
MODEL_DIR = os.path.join(BASE_DIR, "models")
MODEL_PATH = os.path.join(MODEL_DIR, "resume_model.pkl")
JOB_INDEX_PATH = os.path.join(MODEL_DIR, "job_index.json")

app = Flask(
    __name__,
    static_folder=STATIC_DIR,
    template_folder=TEMPLATES_DIR
)
CORS(app)
bcrypt = Bcrypt(app)

# ---------------- MongoDB (for users only) ----------------
MONGO_URI = "mongodb://localhost:27017/"
client = MongoClient(MONGO_URI)
db = client['jobsinline']
users_collection = db['users']

# ---------------- Gemini API Key (Hardcoded) ----------------
GOOGLE_API_KEY = "AIzaSyCCHFgeeK7ToNo4nQ6PivPsJB4IakqHxj4"
GOOGLE_API_URL = (
    f'https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={GOOGLE_API_KEY}'
)

# ---------------- Utilities ----------------
def load_job_roles_from_excel():
    if not os.path.exists(EXCEL_FILE_PATH):
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() for h in rows[0]]
    job_roles = []
    for row in rows[1:]:
        job = dict(zip(headers, row))
        job_roles.append(job)
    return job_roles

def save_job_roles_to_excel(job_roles):
    if not job_roles:
        return
    headers = list(job_roles[0].keys())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for job in job_roles:
        ws.append([job.get(h, "") for h in headers])
    wb.save(EXCEL_FILE_PATH)

def find_job_in_excel(job_role):
    job_roles = load_job_roles_from_excel()
    for job in job_roles:
        if str(job.get("JOB ROLES", "")).strip().lower() == job_role.strip().lower():
            return job
    return None

def safe_pdf_to_text(filepath: str) -> str:
    text = ""
    with open(filepath, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        for i, page in enumerate(reader.pages):
            try:
                page_text = page.extract_text() or ""
            except Exception as e:
                logging.warning(f"PDF extract warning page {i}: {e}")
                page_text = ""
            text += page_text + "\n"
    return text.strip()

def perform_analysis(resume_text, required_skills_str, required_frameworks_str):
    # Support both string and list for skills/frameworks
    if isinstance(required_skills_str, list):
        required_skills = [s.strip() for s in required_skills_str if s and str(s).strip()]
    else:
        required_skills = [s.strip() for s in str(required_skills_str or '').split(',') if s.strip()]

    if isinstance(required_frameworks_str, list):
        required_frameworks = [f.strip() for f in required_frameworks_str if f and str(f).strip()]
    else:
        required_frameworks = [f.strip() for f in str(required_frameworks_str or '').split(',') if f.strip()]

    lt = resume_text.lower()
    skills_found = [s for s in required_skills if s.lower() in lt]
    frameworks_found = [f for f in required_frameworks if f.lower() in lt]

    additional_skills = [s for s in required_skills if s not in skills_found]
    additional_frameworks = [f for f in required_frameworks if f not in frameworks_found]

    skills_prob = (len(skills_found) / len(required_skills) * 50) if required_skills else 0
    frameworks_prob = (len(frameworks_found) / len(required_frameworks) * 50) if required_frameworks else 0
    probability = round(skills_prob + frameworks_prob)

    if probability == 100:
        feedback = 'Great job! You are a perfect match for this role!'
    elif probability >= 50:
        feedback = 'You have some required skills. To improve, focus on: ' + ", ".join(additional_skills + additional_frameworks)
    else:
        feedback = 'Significant improvement needed. You should learn: ' + ", ".join(additional_skills + additional_frameworks)

    return {
        'probability': int(probability),
        'additionalSkills': ', '.join(additional_skills) or 'None',
        'additionalFrameworks': ', '.join(additional_frameworks) or 'None',
        'feedback': feedback,
    }

def get_job_requirements_from_ai(job_role):
    prompt = (
        "Provide the basic programming languages (skills) and frameworks required for the job role "
        f"'{job_role}'. Respond in this format:\nSkills: skill1, skill2\nFrameworks: framework1, framework2"
    )
    try:
        r = requests.post(GOOGLE_API_URL, json={"contents": [{"parts": [{"text": prompt}]}]}, timeout=12)
        logging.info(f"AI API status: {r.status_code}")
        if r.status_code != 200:
            logging.error(f"Gemini non-200: {r.text[:300]}")
            return '', ''
        data = r.json()
        txt = data['candidates'][0]['content']['parts'][0]['text']

        skills_match = re.search(r'Skills:\s*(.+)', txt, re.I)
        frameworks_match = re.search(r'Frameworks:\s*(.+)', txt, re.I)
        required_skills = skills_match.group(1).strip() if skills_match else ''
        required_frameworks = frameworks_match.group(1).strip() if frameworks_match else ''
        return required_skills, required_frameworks
    except Exception as e:
        logging.error(f"AI error: {e}")
        return '', ''

# ---------------- ML: training + inference ----------------
def build_job_corpus_from_excel():
    corpus = []
    job_roles = load_job_roles_from_excel()
    for rec in job_roles:
        role = rec.get('JOB ROLES', '') or ''
        skills = rec.get('PROGRAMMING SKILLS', '') or ''
        frameworks = rec.get('FRAMEWORKS', '') or ''
        blob = " ".join([str(role), str(skills), str(frameworks)])
        corpus.append({
            "role": role,
            "text": blob
        })
    return corpus

def train_model_from_excel(force: bool = False):
    os.makedirs(MODEL_DIR, exist_ok=True)
    if os.path.exists(MODEL_PATH) and not force:
        return load_model()
    corpus = build_job_corpus_from_excel()
    if not corpus:
        logging.warning("No job roles found in Excel to train model.")
        return None
    X = [c['text'] for c in corpus]
    y = [str(c['role']).strip().lower() for c in corpus]
    pipe = Pipeline([
        ("tfidf", TfidfVectorizer(ngram_range=(1,2), stop_words='english', min_df=1)),
        ("clf", LogisticRegression(max_iter=1000, n_jobs=None))
    ])
    pipe.fit(X, y)
    joblib.dump(pipe, MODEL_PATH)
    with open(JOB_INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(corpus, f, ensure_ascii=False, indent=2)
    logging.info(f"ML model trained and saved to {MODEL_PATH} (classes: {len(set(y))})")
    return pipe

def load_model():
    if not os.path.exists(MODEL_PATH):
        return None
    try:
        pipe = joblib.load(MODEL_PATH)
        return pipe
    except Exception as e:
        logging.error(f"Failed loading model: {e}")
        return None

ML_PIPELINE = train_model_from_excel(force=False)

def ensure_model():
    global ML_PIPELINE
    if ML_PIPELINE is None:
        ML_PIPELINE = train_model_from_excel(force=True)
    return ML_PIPELINE

def top_similar_jobs(resume_text, top_k=10):
    ensure_model()
    if ML_PIPELINE is None or not os.path.exists(JOB_INDEX_PATH):
        jobs = load_job_roles_from_excel()
        results = []
        lt = resume_text.lower()
        for j in jobs:
            sk = j.get('PROGRAMMING SKILLS', '') or ''
            fw = j.get('FRAMEWORKS', '') or ''
            analysis = perform_analysis(lt, sk, fw)
            results.append({
                "jobRole": j.get('JOB ROLES', 'Unknown Job'),
                "score": analysis['probability'] / 100.0
            })
        results.sort(key=lambda x: x["score"], reverse=True)
        return results[:top_k]

    with open(JOB_INDEX_PATH, "r", encoding="utf-8") as f:
        corpus = json.load(f)

    tfidf = ML_PIPELINE.named_steps["tfidf"]
    r_vec = tfidf.transform([resume_text])
    j_matrix = tfidf.transform([c["text"] for c in corpus])
    sims = cosine_similarity(r_vec, j_matrix).ravel()

    ranked = sorted(
        [{"jobRole": corpus[i]["role"], "score": float(sims[i])} for i in range(len(corpus))],
        key=lambda x: x["score"], reverse=True
    )
    return ranked[:top_k]

def ml_predict_role(resume_text):
    ensure_model()
    if ML_PIPELINE is None:
        return None
    clf = ML_PIPELINE.named_steps["clf"]
    pred = ML_PIPELINE.predict([resume_text])[0]
    if hasattr(clf, "predict_proba"):
        proba = ML_PIPELINE.predict_proba([resume_text])[0]
        classes = ML_PIPELINE.classes_
        idx = list(classes).index(pred)
        conf = float(proba[idx])
    else:
        conf = None
    return {"standardizedRole": pred, "confidence": conf}

# ---------------- Routes ----------------
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/main.html')
def main_page():
    return render_template('main.html')

@app.route('/auth', methods=['POST'])
def auth():
    req = request.get_json(force=True)
    action = req.get('action')
    username = req.get('username')
    password = req.get('password')

    if action == 'login':
        user = users_collection.find_one({'Username': username})
        if not user:
            return jsonify({'success': False, 'message': 'No username found! Please register.', 'redirect': 'register'}), 404
        if not bcrypt.check_password_hash(user['Password'], password):
            return jsonify({'success': False, 'message': 'Incorrect password!'}), 401
        return jsonify({'success': True, 'message': 'Login successful!', 'redirect': 'main', 'username': username})

    elif action == 'register':
        email = req.get('email')
        phone = req.get('phone')
        if users_collection.find_one({'Username': username}):
            return jsonify({'success': False, 'message': 'Username already exists!'}), 400
        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
        users_collection.insert_one({'Username': username, 'Email': email, 'Phone': phone, 'Password': hashed_password})
        return jsonify({'success': True, 'message': 'Registration successful! You can now log in.'}), 201

    return jsonify({'success': False, 'message': 'Invalid action'}), 400

@app.route('/upload', methods=['POST'])
def upload():
    if 'resume' not in request.files or 'jobRole' not in request.form:
        return jsonify({'success': False, 'error': 'No file or job role provided'}), 400

    file = request.files['resume']
    job_role = request.form['jobRole']

    if file.filename == '':
        return jsonify({'success': False, 'error': 'No selected file'}), 400
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'error': 'Only PDF files are allowed'}), 400

    os.makedirs('uploads', exist_ok=True)
    file_path = os.path.join('uploads', file.filename)
    file.save(file_path)

    try:
        resume_text = safe_pdf_to_text(file_path)
        if not resume_text:
            logging.error("PDF extraction failed: No text found in file.")
            return jsonify({'success': False, 'error': 'Could not extract text from PDF.'}), 400

        job_data = find_job_in_excel(job_role)
        analysis_result = None
        fromChatbot = False

        if job_data:
            analysis_result = perform_analysis(
                resume_text,
                job_data.get('PROGRAMMING SKILLS'),
                job_data.get('FRAMEWORKS')
            )
        else:
            req_skills, req_fw = get_job_requirements_from_ai(job_role)
            analysis_result = perform_analysis(resume_text, req_skills, req_fw)
            fromChatbot = True
            # Save new job role only if AI analysis succeeds (probability ≠ 0)
            if analysis_result['probability'] != 0:
                job_roles = load_job_roles_from_excel()
                job_roles.append({
                    'JOB ROLES': job_role,
                    'PROGRAMMING SKILLS': req_skills,
                    'FRAMEWORKS': req_fw
                })
                save_job_roles_to_excel(job_roles)
                train_model_from_excel(force=True)  # retrain ML model

        ml_pred = ml_predict_role(resume_text)
        related = top_similar_jobs(resume_text, top_k=8)
        requested_match = None
        if related:
            for r in related:
                if str(r["jobRole"]).strip().lower() == job_role.strip().lower():
                    requested_match = r["score"]
                    break

        response = {
            'success': True,
            'jobRole': job_role,
            'probability': analysis_result['probability'],
            'additionalSkills': analysis_result['additionalSkills'],
            'additionalFrameworks': analysis_result['additionalFrameworks'],
            'feedback': analysis_result['feedback'],
            'fromChatbot': fromChatbot,
            'mlPredictedRole': (ml_pred or {}).get('standardizedRole'),
            'mlConfidence': (ml_pred or {}).get('confidence'),
            'requestedRoleSimilarity': None if requested_match is None else round(requested_match * 100),
            'topRelatedJobs': [
                {'jobRole': r['jobRole'], 'similarity': round(r['score'] * 100)}
                for r in related
            ]
        }
        return jsonify(response)

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logging.error(f"Error processing resume analysis: {e}\n{error_details}")
        return jsonify({
            'success': False,
            'error': 'Error processing resume analysis',
            'details': str(e),
            'traceback': error_details
        }), 500
    finally:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass

@app.route('/related_jobs', methods=['POST'])
def related_jobs():
    if 'resume' not in request.files:
        return jsonify({'success': False, 'error': 'No resume file provided'}), 400

    file = request.files['resume']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No selected file'}), 400
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'error': 'Only PDF files are allowed'}), 400

    os.makedirs('uploads', exist_ok=True)
    path = os.path.join('uploads', file.filename)
    file.save(path)

    try:
        resume_text = safe_pdf_to_text(path)
        if not resume_text:
            return jsonify({'success': False, 'error': 'Could not extract text from PDF.'}), 400

        sims = top_similar_jobs(resume_text, top_k=20)
        related_jobs_list = [
            {
                'jobRole': r['jobRole'],
                'probability': int(round(r['score'] * 100)),
                'additionalSkills': '—',
                'additionalFrameworks': '—'
            } for r in sims
        ]
        return jsonify({'success': True, 'relatedJobs': related_jobs_list})

    except Exception as e:
        logging.error(f"Unexpected error in related_jobs: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'An unexpected error occurred while finding related jobs.', 'details': str(e)}), 500
    finally:
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass

@app.route('/chatbot', methods=['POST'])
def chatbot():
    req = request.get_json(force=True)
    message = req.get('message')
    if not message:
        return jsonify({'success': False, 'error': 'No message provided'}), 400

    if not GOOGLE_API_URL:
        return jsonify({'success': True, 'response': f"I received: {message.strip()}. How can I help with your resume or job search?"})

    try:
        prompt = (
            "Answer succinctly and clearly. Fix punctuation if needed.\n"
            f"User: {message}\nBot:"
        )
        r = requests.post(GOOGLE_API_URL, json={"contents": [{"parts": [{"text": prompt}]}]}, timeout=10)
        if r.status_code != 200:
            logging.error(f"Chatbot AI service error: {r.text[:200]}")
            return jsonify({'success': True, 'response': "I'm having trouble connecting to my AI service right now. Please try again later."})

        response_text = r.json()['candidates'][0]['content']['parts'][0]['text']
        response_text = re.sub(r'\s+', ' ', response_text).strip()
        response_text = re.sub(r'\s+([.,!?;:])', r'\1', response_text)
        short = response_text.split('. ')[0].strip()
        if not short.endswith('.'):
            short += '.'
        return jsonify({'success': True, 'response': short})

    except Exception as e:
        logging.error(f"Chatbot exception: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Error processing chatbot request'}), 500

# ---------------- Main ----------------
if __name__ == '__main__':
    if not os.path.exists(MODEL_PATH):
        logging.info("No ML model found. Training now from Excel data...")
        train_model_from_excel(force=True)
    app.run(host='0.0.0.0', port=5000, debug=True)
